import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def load_evaluation_sheet(file_path, sheet_name="EVALUACIÓN"):
    """Carga la hoja de evaluación y devuelve el DataFrame."""
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    return xls.parse(sheet_name)


def find_student_row(df, student_name):
    """Busca la fila donde se encuentra el estudiante."""
    for index, row in df.iterrows():
        if str(row.iloc[2]).strip().lower() == student_name.lower():
            return index
    return None


def find_activity_column(df, activity_name):
    """Busca la columna donde está la actividad."""
    activity_row = 7  # La fila 7 contiene los nombres de actividades
    for col_index, value in enumerate(df.iloc[activity_row]):
        if str(value).strip().lower() == activity_name.lower():
            return col_index
    return None


def get_cell_value(sheet, cell_reference):
    """Obtiene el valor de una celda específica (Ej: 'AC10')."""
    try:
        return sheet[cell_reference].value
    except KeyError:
        return "Celda no encontrada"


def locate_grade_cell(file_path, student_name, activity_name):
    """
    Devuelve la referencia de la celda donde se debe insertar la nota
    y el valor actual contenido en esa celda.
    """
    df = load_evaluation_sheet(file_path)
    student_row = find_student_row(df, student_name)
    activity_column = find_activity_column(df, activity_name)

    if student_row is None:
        return f"Estudiante '{student_name}' no encontrado."
    if activity_column is None:
        return f"Actividad '{activity_name}' no encontrada."

    # Se suma 1 a ambos índices para obtener la referencia en Excel (base 1)
    column_letter = get_column_letter(activity_column + 1)
    cell_reference = f"{column_letter}{student_row + 1}"

    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb["EVALUACIÓN"]

    cell_value = get_cell_value(sheet, cell_reference)

    return f"{cell_reference} (Valor actual: {cell_value})"


def browse_file():
    """Permite seleccionar el archivo Excel y lo muestra en el entry correspondiente."""
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry_file_path.delete(0, tk.END)
    entry_file_path.insert(0, file_path)


def search_cell():
    """Utiliza los datos ingresados para buscar la celda correspondiente en el Excel."""
    file_path = entry_file_path.get()
    student_name = entry_student.get()
    activity_name = entry_activity.get()
    result = locate_grade_cell(file_path, student_name, activity_name)
    label_result.config(text=result)


def run_tests():
    """
    Ejecuta casos de prueba predefinidos para verificar el mapeo de celdas.
    Cada caso de prueba contiene:
      - Nombre del alumno.
      - Nombre de la actividad.
      - Valor esperado (por ejemplo, la referencia de celda esperada).
    Los resultados se muestran en una ventana emergente.
    """
    # Definir casos de prueba (ajusta estos datos de acuerdo al Excel de pruebas)
    test_cases = [
        {"student": "González González, Marta", "activity": "1K", "expected": "6.3"},
        {"student": "Principe Álvarez, Mirko", "activity": "1H", "expected": "6"}
    ]

    file_path = entry_file_path.get()
    if not file_path:
        messagebox.showerror("Error", "Por favor, seleccione un archivo Excel primero.")
        return

    results = []
    for test in test_cases:
        result = locate_grade_cell(file_path, test["student"], test["activity"])
        # Se puede comparar la celda obtenida con el valor esperado (en este ejemplo se muestra la comparación)
        results.append(
            f"Alumno: {test['student']} | Actividad: {test['activity']} -> "
            f"{result} | Esperado: {test['expected']}"
        )

    # Mostrar los resultados en una ventana emergente
    test_window = tk.Toplevel(root)
    test_window.title("Resultados de Pruebas de Mapeo")
    text_widget = tk.Text(test_window, width=80, height=10)
    text_widget.pack(padx=10, pady=10, fill="both", expand=True)
    for line in results:
        text_widget.insert(tk.END, line + "\n")
    text_widget.config(state="disabled")


# Configuración de la interfaz principal
root = tk.Tk()
root.title("Verificación de Mapeo de Calificaciones")

# Selección del archivo Excel
tk.Label(root, text="Ruta del archivo Excel:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_file_path = tk.Entry(root, width=50)
entry_file_path.grid(row=0, column=1, padx=5, pady=5)
tk.Button(root, text="Buscar", command=browse_file).grid(row=0, column=2, padx=5, pady=5)

# Entrada para nombre del estudiante
tk.Label(root, text="Nombre del Estudiante:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_student = tk.Entry(root, width=30)
entry_student.grid(row=1, column=1, padx=5, pady=5, sticky="w")

# Entrada para nombre de la actividad
tk.Label(root, text="Nombre de la Actividad:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_activity = tk.Entry(root, width=30)
entry_activity.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# Botón para buscar la celda correspondiente
tk.Button(root, text="Buscar Celda", command=search_cell).grid(row=3, column=1, padx=5, pady=5, sticky="w")
label_result = tk.Label(root, text="", fg="blue")
label_result.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

# Botón para ejecutar casos de prueba de mapeo
tk.Button(root, text="Ejecutar Pruebas de Mapeo", command=run_tests).grid(row=5, column=1, padx=5, pady=5, sticky="w")

root.mainloop()
