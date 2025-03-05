import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

WORKBOOK_CACHE = {}


def load_evaluation_sheet(file_path, sheet_name="EVALUACIÓN"):
    """Carga la hoja de evaluación con caché."""
    if file_path in WORKBOOK_CACHE:
        wb, df = WORKBOOK_CACHE[file_path]
    else:
        try:
            wb = load_workbook(filename=file_path, data_only=True)
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            WORKBOOK_CACHE[file_path] = (wb, df)
        except Exception as e:
            raise Exception(f"Error al cargar el archivo: {str(e)}")
    return wb, df


def find_student_row(df, student_name):
    """Busca la fila donde se encuentra el estudiante."""
    clean_names = df.iloc[:, 2].astype(str).str.strip().str.lower()
    target = student_name.strip().lower()

    matches = clean_names[clean_names == target]
    if matches.empty:
        raise Exception(f"Estudiante '{student_name}' no encontrado.")

    return matches.index[0]


def find_activity_column(df, activity_name):
    """Busca la columna donde está la actividad."""
    activity_row = df.iloc[7].astype(str).str.strip().str.lower()
    target = activity_name.strip().lower()

    exact_match = activity_row[activity_row == target]
    if not exact_match.empty:
        return exact_match.index[0]

    raise Exception(f"Actividad '{activity_name}' no encontrada.")


def locate_grade_cell(file_path, student_name, activity_name):
    """Devuelve la referencia de celda y su valor."""
    try:
        wb, df = load_evaluation_sheet(file_path)
        student_row = find_student_row(df, student_name)
        activity_column = find_activity_column(df, activity_name)

        sheet = wb["EVALUACIÓN"]
        column_letter = get_column_letter(activity_column + 1)
        cell_reference = f"{column_letter}{student_row + 1}"

        cell_value = sheet[cell_reference].value

        # Convertir el valor de la celda a cadena para evitar errores de concatenación
        cell_value_str = "Celda vacía" if cell_value is None else str(cell_value)

        return f"{cell_reference} (Valor actual: {cell_value_str})"
    except Exception as e:
        return f"Error: {str(e)}"


class GradingGUI:
    def __init__(self, master):
        self.master = master
        master.title("Gestión de Calificaciones v2.0")

        self.frame = tk.Frame(master, padx=20, pady=20)
        self.frame.pack(expand=True, fill="both")

        tk.Label(self.frame, text="Archivo Excel:").grid(row=0, column=0, sticky="w")
        self.entry_file = tk.Entry(self.frame, width=40)
        self.entry_file.grid(row=0, column=1)

        tk.Button(self.frame, text="Examinar", command=self.load_file).grid(row=0, column=2)

        tk.Label(self.frame, text="Estudiante (Apellido, Nombre):").grid(row=1, column=0)
        self.entry_student = tk.Entry(self.frame)
        self.entry_student.grid(row=1, column=1)

        tk.Label(self.frame, text="Actividad:").grid(row=2, column=0)
        self.entry_activity = tk.Entry(self.frame)
        self.entry_activity.grid(row=2, column=1)

        tk.Button(self.frame, text="Buscar Celda", command=self.search_cell).grid(row=3, column=1)

        self.lbl_result = tk.Label(self.frame, text="", fg="blue")
        self.lbl_result.grid(row=4, columnspan=3)

    def load_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, path)

    def search_cell(self):
        try:
            file_path = self.entry_file.get()
            student_name = self.entry_student.get()
            activity_name = self.entry_activity.get()

            result = locate_grade_cell(file_path, student_name.strip(), activity_name.strip())
            self.lbl_result.config(text=result)

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.lbl_result.config(text="", fg="red")


if __name__ == "__main__":
    root = tk.Tk()
    app = GradingGUI(root)
    root.mainloop()
