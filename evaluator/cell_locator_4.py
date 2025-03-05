import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def build_activity_mapping(file_path,
                           sheet_name="EVALUACIÓN",
                           header_text="RESULTADO APRENDIZAJE-CRITERIO DE EVALUACIÓN PRÁCTICOS",
                           activity_row=9,
                           student_col="C",
                           start_student_row=10):
    """
    Crea un mapeo que asocia cada 'TAREA X' a la(s) columna(s) donde aparece.
    Retorna:
      - mapping: Diccionario del tipo {"TAREA 1": ["D9", "L9", "S9"], ...}
      - find_student_row: Función que, dado el nombre de un alumno, retorna la fila en la que se encuentra.
      - get_grade_cell: Función que, dada una tarea, un alumno y un índice de trimestre, retorna la referencia de la celda donde se debe insertar la nota.
    """
    # Cargar el workbook en modo data_only para obtener los valores (resultado de fórmulas)
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # 1. Detectar bloques de celdas combinadas que contengan el header_text
    merged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        if top_left_value and header_text.lower() in str(top_left_value).lower():
            merged_ranges.append(merged_range)

    # 2. Recorrer cada bloque para identificar las columnas en la fila de actividades que contengan "TAREA [X]"
    mapping = {}
    for rng in merged_ranges:
        for col in range(rng.min_col, rng.max_col + 1):
            cell_value = ws.cell(row=activity_row, column=col).value
            if cell_value and re.search(r"tarea\s*\d+", str(cell_value), re.IGNORECASE):
                task_name = str(cell_value).strip()
                col_letter = get_column_letter(col)
                if task_name not in mapping:
                    mapping[task_name] = []
                mapping[task_name].append(f"{col_letter}{activity_row}")

    # 3. Función interna para encontrar la fila de un alumno (se asume que los nombres están en la columna 'student_col')
    def find_student_row(alumno):
        print(f"[DEBUG] Buscando alumno: '{alumno}'")
        for row in range(start_student_row, ws.max_row + 1):
            cell_value = ws[f"{student_col}{row}"].value
            print(f"[DEBUG] Fila {row}, valor: {cell_value}")
            if cell_value and alumno.lower() in str(cell_value).lower():
                print(f"[DEBUG] Alumno encontrado en la fila {row}")
                return row
        print("[DEBUG] Alumno no encontrado")
        return None

    # 4. Función para obtener la referencia de la celda donde se colocará la nota,
    #    considerando el índice de trimestre (0 para el primero, 1 para el segundo, etc.)
    def get_grade_cell(alumno, tarea, trimester_index=0):
        if tarea not in mapping:
            return None
        # Si hay varias celdas para la misma tarea, usamos la correspondiente al índice
        cell_refs = mapping[tarea]
        if len(cell_refs) > trimester_index:
            selected_ref = cell_refs[trimester_index]
        else:
            selected_ref = cell_refs[0]  # Si no existe el índice, se usa la primera
        col_letters = "".join([ch for ch in selected_ref if ch.isalpha()])
        row_alumno = find_student_row(alumno)
        if not row_alumno:
            return None
        return f"{col_letters}{row_alumno}"

    return mapping, find_student_row, get_grade_cell


# Interfaz gráfica con Tkinter
class GradingGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestión de Calificaciones")
        self.geometry("650x450")
        self.file_path = None
        self.mapping = None
        self.find_student_row = None
        self.get_grade_cell = None

        self.create_widgets()

    def create_widgets(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)

        # Selección de archivo
        ttk.Label(frame, text="Archivo Excel:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.entry_file = ttk.Entry(frame, width=50)
        self.entry_file.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame, text="Examinar", command=self.load_file).grid(row=0, column=2, padx=5, pady=5)

        # Entrada para nombre del alumno
        ttk.Label(frame, text="Nombre del Alumno:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.entry_student = ttk.Entry(frame, width=40)
        self.entry_student.grid(row=1, column=1, padx=5, pady=5)

        # Entrada para la tarea
        ttk.Label(frame, text="Tarea (ej. TAREA 1):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.entry_task = ttk.Entry(frame, width=40)
        self.entry_task.grid(row=2, column=1, padx=5, pady=5)

        # Selector de trimestre
        ttk.Label(frame, text="Trimestre:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.combo_trimester = ttk.Combobox(frame, values=["1T", "2T", "3T"], state="readonly", width=10)
        self.combo_trimester.current(0)  # Por defecto 1T
        self.combo_trimester.grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)

        # Botón para buscar la celda de la nota
        ttk.Button(frame, text="Buscar Celda", command=self.search_cell).grid(row=4, column=1, pady=10)

        # Área de salida
        self.text_output = tk.Text(frame, height=10, width=70)
        self.text_output.grid(row=5, column=0, columnspan=3, padx=5, pady=5)

    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.file_path = file_path
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, file_path)
            try:
                # Construir el mapeo y obtener las funciones auxiliares
                self.mapping, self.find_student_row, self.get_grade_cell = build_activity_mapping(self.file_path)
                messagebox.showinfo("Archivo cargado", "Archivo cargado y mapeo generado correctamente.")
                self.text_output.delete(1.0, tk.END)
                self.text_output.insert(tk.END, "Mapeo de Tareas:\n")
                for tarea, refs in self.mapping.items():
                    self.text_output.insert(tk.END, f"  {tarea} -> {refs}\n")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def search_cell(self):
        if not self.file_path:
            messagebox.showwarning("Advertencia", "Primero debe cargar un archivo Excel.")
            return
        alumno = self.entry_student.get().strip()
        tarea = self.entry_task.get().strip()
        trimester_str = self.combo_trimester.get().strip()
        if not alumno or not tarea:
            messagebox.showwarning("Advertencia", "Debe ingresar el nombre del alumno y la tarea.")
            return

        # Convertir el trimestre a índice (1T -> 0, 2T -> 1, 3T -> 2)
        trimester_index = 0
        if trimester_str.upper() == "2T":
            trimester_index = 1
        elif trimester_str.upper() == "3T":
            trimester_index = 2

        cell_ref = self.get_grade_cell(alumno, tarea, trimester_index)
        if cell_ref:
            # Para obtener el valor de la celda, cargar el workbook en modo data_only
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb["EVALUACIÓN"]
            grade_value = ws[cell_ref].value
            self.text_output.insert(tk.END, f"\nCelda para '{alumno}' y '{tarea}' ({trimester_str}): {cell_ref}")
            self.text_output.insert(tk.END, f"\nValor (nota): {grade_value}\n")
        else:
            messagebox.showerror("Error", f"No se pudo determinar la celda para '{alumno}' y '{tarea}'.")

if __name__ == "__main__":
    app = GradingGUI()
    app.mainloop()
