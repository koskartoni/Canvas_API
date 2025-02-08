import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- Funciones de mapeo de actividades ---

def build_activity_mapping(file_path, sheet_name="EVALUACIÓN",
                           merged_header_text="RESULTADO APRENDIZAJE-CRITERIO DE EVALUACIÓN PRÁCTICOS",
                           header_rows=(7, 8), activity_row=9):
    """
    Abre el archivo Excel y recorre los rangos combinados en las filas de encabezado (header_rows)
    que contienen el texto especificado en merged_header_text.
    Para cada bloque encontrado, extrae de la fila activity_row los nombres de las actividades
    y construye un diccionario en el que cada actividad se asocia a una lista de referencias de celda.
    Por ejemplo: { "1 h": ["AC9", "AE9"], "1K": ["AD9"] }
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    mapping = {}

    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # Verificamos que el rango abarque las filas de encabezado definidas (7 y 8)
        if min_row == header_rows[0] and max_row == header_rows[1]:
            header_value = ws.cell(row=min_row, column=min_col).value
            if header_value and merged_header_text.lower() in str(header_value).lower():
                # Se ha identificado un bloque de actividades; recorremos cada columna de este bloque
                for col in range(min_col, max_col + 1):
                    cell = ws.cell(row=activity_row, column=col)
                    if cell.value is not None:
                        activity_name = str(cell.value).strip()
                        cell_ref = f"{get_column_letter(col)}{activity_row}"
                        # Si la tarea ya existe en el mapeo, se añade la referencia a la lista;
                        # de lo contrario, se crea una nueva entrada.
                        if activity_name in mapping:
                            mapping[activity_name].append(cell_ref)
                        else:
                            mapping[activity_name] = [cell_ref]
    return mapping

# --- Función para encontrar la fila del alumno usando pandas ---

def find_student_row(file_path, student_name, sheet_name="EVALUACIÓN"):
    """
    Busca la fila en la que se encuentra el alumno.
    Se asume que los nombres están en la columna C (índice 2, 0-indexado) a partir de la fila 10.
    Devuelve el número de fila (1-indexado) tal como aparece en el Excel.
    """
    try:
        # Leer el Excel sin interpretar filas como encabezado
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", header=None)
    except Exception as e:
        return None

    # Iteramos desde la fila 10 (índice 9 en pandas, 0-indexado)
    for idx in range(9, len(df)):
        cell_value = df.iloc[idx, 2]
        if cell_value is not None and str(cell_value).strip().lower() == student_name.strip().lower():
            # Sumamos 1 para convertir el índice 0-indexado en número de fila real del Excel
            return idx + 1
    return None

# --- Función para verificar la correspondencia alumno-tarea-nota ---

def verify_grade_for_student():
    """
    Para el alumno y la tarea especificados en la interfaz:
      1. Busca la fila del alumno.
      2. Obtiene la(s) columna(s) asociada(s) a la tarea a través del mapeo.
      3. Construye la referencia de celda (reemplazando la fila del header por la del alumno).
      4. Lee el valor actual de esa celda y lo muestra.
    """
    file_path = entry_file_path.get().strip()
    student = entry_student.get().strip()
    task = entry_task.get().strip()

    if not file_path or not student or not task:
        messagebox.showerror("Error", "Por favor, ingrese la ruta del archivo, el alumno y la tarea.")
        return

    mapping = build_activity_mapping(file_path)
    if task not in mapping:
        messagebox.showerror("Error", f"La tarea '{task}' no se encuentra en el mapeo.")
        return

    student_row = find_student_row(file_path, student)
    if student_row is None:
        messagebox.showerror("Error", f"Alumno '{student}' no encontrado.")
        return

    # Abrir el libro para leer el valor de la celda
    wb = load_workbook(file_path, data_only=True)
    ws = wb["EVALUACIÓN"]

    results = f"Verificación para alumno '{student}', tarea '{task}':\n\n"
    # mapping[task] es una lista de celdas de la fila de cabecera (por ejemplo, ["AC9", "AD9"])
    for cell_ref in mapping[task]:
        # Extraemos la parte de la columna (letras) del cell_ref
        col_letter = ''.join([ch for ch in cell_ref if ch.isalpha()])
        # Construimos la celda para el alumno usando la fila encontrada
        student_cell = f"{col_letter}{student_row}"
        cell_value = ws[student_cell].value
        results += f"Celda {student_cell}: Valor actual -> {cell_value}\n"

    messagebox.showinfo("Verificación de Nota", results)

# --- Interfaz Gráfica con Tkinter ---

root = tk.Tk()
root.title("Verificar Correspondencia Alumno-Tarea-Nota")

# Selección del archivo Excel
tk.Label(root, text="Ruta del archivo Excel:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_file_path = tk.Entry(root, width=50)
entry_file_path.grid(row=0, column=1, padx=5, pady=5)
tk.Button(root, text="Seleccionar Archivo", command=lambda:
          entry_file_path.insert(0, filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]))).grid(row=0, column=2, padx=5, pady=5)

# Entrada para el nombre del alumno
tk.Label(root, text="Nombre del Alumno:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_student = tk.Entry(root, width=30)
entry_student.grid(row=1, column=1, padx=5, pady=5, sticky="w")

# Entrada para el nombre de la tarea
tk.Label(root, text="Nombre de la Tarea:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_task = tk.Entry(root, width=30)
entry_task.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# Botón para verificar la nota
tk.Button(root, text="Verificar Nota", command=verify_grade_for_student).grid(row=3, column=1, padx=5, pady=10)

root.mainloop()
